"""
Script to generate an Excel template for importing exam papers.
Uses openpyxl (lightweight, no pandas required).
Run: python generate_exam_template_simple.py
Output: exam_import_template.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import argparse
import os

def generate_exam_template(output_file: str | None = None):
    """Generate an Excel template for importing exam papers

    Args:
        output_file: Optional custom output path. Defaults to 'exam_import_template.xlsx'.
    """
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Exam_Info sheet
    ws_info = wb.create_sheet('Exam_Info')
    
    # Header style
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Exam Info headers
    info_headers = ['Exam Name', 'Description', 'Duration (minutes)', 'Total Marks', 'Language']
    ws_info.append(info_headers)
    
    # Style headers
    for cell in ws_info[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Example data
    ws_info.append([
        'Mathematics Midterm Exam',
        'Midterm examination covering algebra and calculus topics',
        90,
        100,
        'English'
    ])
    
    # Auto-adjust column widths
    for column in ws_info.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_info.column_dimensions[column_letter].width = adjusted_width
    
    # Create Test Paper 1 sheet
    ws_paper1 = wb.create_sheet('PAPER001')
    
    # Question columns
    question_headers = [
        'Question Type', 'Question Number', 'Question Text', 'Question Image',
        'Option A', 'Option B', 'Option C', 'Option D',
        'Correct Option', 'Correct Options', 'Correct Answer',
        'Marks', 'Negative Marks', 'Topic'
    ]
    ws_paper1.append(question_headers)
    
    # Style headers
    for cell in ws_paper1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Example MCQ question
    ws_paper1.append([
        'MCQ',  # Question Type
        1,      # Question Number
        'What is the derivative of x^2?',  # Question Text
        '',     # Question Image (optional)
        '2x',   # Option A
        'x',    # Option B
        '2',    # Option C
        'x^2',  # Option D
        'A',    # Correct Option (MCQ)
        '',     # Correct Options (MSQ - leave empty)
        '',     # Correct Answer (Numerical - leave empty)
        2.0,    # Marks
        0.5,    # Negative Marks
        'Calculus'  # Topic
    ])
    
    # Example MSQ question
    ws_paper1.append([
        'MSQ',
        2,
        'Which of the following are prime numbers?',
        '',
        '2',
        '3',
        '4',
        '5',
        '',     # Correct Option (leave empty for MSQ)
        'ABD',  # Correct Options (MSQ: can be AB, ABC, ABCD, etc.)
        '',     # Correct Answer (leave empty)
        3.0,
        1.0,
        'Number Theory'
    ])
    
    # Example Numerical question
    ws_paper1.append([
        'Numerical',
        3,
        'Solve for x: 2x + 5 = 13',
        '',
        '',     # Option A (leave empty for Numerical)
        '',     # Option B
        '',     # Option C
        '',     # Option D
        '',     # Correct Option (leave empty)
        '',     # Correct Options (leave empty)
        4.0,    # Correct Answer (Numerical: numeric value)
        2.0,
        0.5,
        'Algebra'
    ])
    
    # Auto-adjust column widths for paper sheet
    for column in ws_paper1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_paper1.column_dimensions[column_letter].width = adjusted_width
    
    # Create Test Paper 2 sheet (with fewer examples)
    ws_paper2 = wb.create_sheet('PAPER002')
    ws_paper2.append(question_headers)
    
    # Style headers
    for cell in ws_paper2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Example questions for paper 2
    ws_paper2.append([
        'MCQ',
        1,
        'What is the integral of 1/x?',
        '',
        'x',
        'ln(x)',
        '1/x',
        'x^2',
        'B',
        '',
        '',
        2.0,
        0.5,
        'Calculus'
    ])
    
    ws_paper2.append([
        'Numerical',
        2,
        'Find the value of log(100)',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        2.0,
        1.0,
        0.25,
        'Logarithms'
    ])
    
    # Auto-adjust column widths
    for column in ws_paper2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_paper2.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    if not output_file:
        output_file = 'exam_import_template.xlsx'
    try:
        wb.save(output_file)
    except PermissionError:
        # Likely locked by another program or path not writable. Fallback to timestamped filename.
        base, ext = os.path.splitext(output_file)
        fallback = f"{base}-{datetime.now().strftime('%Y%m%d-%H%M%S')}{ext or '.xlsx'}"
        wb.save(fallback)
        output_file = fallback
    
    print(f"✓ Excel template generated successfully: {output_file}")
    print("\nTemplate Structure:")
    print("  - Sheet 1: 'Exam_Info' - Contains exam details (one row)")
    print("  - Sheet 2+: Test paper sheets with sample questions")
    print("\nColumn Guidelines:")
    print("  - Question Type: Must be 'MCQ', 'MSQ', or 'Numerical'")
    print("  - MCQ: Fill all options (A-D), use 'Correct Option' (single letter)")
    print("  - MSQ: Fill all options (A-D), use 'Correct Options' (combination like 'AB', 'ABC')")
    print("  - Numerical: Leave options empty, use 'Correct Answer' (numeric value)")
    print("\nSee EXAM_IMPORT_GUIDE.md for detailed instructions.")

if __name__ == '__main__':
    try:
        from openpyxl import Workbook
        parser = argparse.ArgumentParser(description='Generate exam import Excel template')
        parser.add_argument('--out', dest='out', default=None, help='Output .xlsx path (optional)')
        args = parser.parse_args()
        generate_exam_template(args.out)
    except ImportError:
        print("Error: Please install openpyxl:")
        print("  pip install openpyxl")
    except Exception as e:
        print(f"Error generating template: {e}")
        import traceback
        traceback.print_exc()

